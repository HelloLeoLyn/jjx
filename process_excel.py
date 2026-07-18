import openpyxl

# Load the workbook
wb = openpyxl.load_workbook(r'C:\Users\leo\Documents\材料.xlsx')
ws = wb.active

print('Sheet:', ws.title)
print('Rows:', ws.max_row, 'Cols:', ws.max_column)

# Print header row
headers = [cell.value for cell in ws[1]]
print('Headers:', headers)

# Print first 5 data rows
for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
    print(list(row))
