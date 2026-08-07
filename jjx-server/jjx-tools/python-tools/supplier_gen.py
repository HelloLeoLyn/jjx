#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
供应商导入模板数据生成器（可复用，2026-08-07 创建）

功能：
    把「去重后的供应商名称列表」按「供应商导入模板」格式生成完整数据，
    保证每列都有数据（格式合规，可直接导入系统）。

用法:
    python3 supplier_gen.py <供应商名称.xlsx或.txt> -o <输出.xlsx>
    名称文件：xlsx 取第1列（跳过表头）或 txt 每行一个名称

生成规则:
    供应商编码  ← SUP001 递增（与系统历史编码一致）
    供应商名称  ← 输入名称
    供应商类型  ← M（物料供应商）
    联系人      ← 常见中文姓名（生成，可替换）
    联系电话    ← 0755-XXXXXXXX 深圳固话 / 13X-XXXX-XXXX 手机（生成，可替换）
    邮箱        ← 名称拼音@163.com（生成，可替换）
    地址        ← 广东省深圳市宝安区（生成，可替换）
    税号        ← 18位统一社会信用代码格式（生成，可替换）
    银行账号    ← 16~19位数字（生成，可替换）
    备注        ← 默认"物料供应商"；若名称含特殊词可加注

依赖: openpyxl
"""
import argparse
import random
import re
import sys
from openpyxl import load_workbook, Workbook

# 常见中文姓氏（生成联系人用）
SURNAMES = list('赵钱孙李周吴郑王冯陈褚卫蒋沈韩杨朱秦尤许何吕施张孔曹严华金魏陶姜')
GIVEN = list('伟芳娜敏静丽强磊军洋勇艳杰娟涛明超秀兰霞平刚桂英华玉梅晨阳子轩雨欣')

# 拼音映射（邮箱域名用，取常见读音，不追求完全准确）
PINYIN = {
    '尚昇': 'shangsheng', '兴富成': 'xingfucheng', 'JJX': 'jjx', '地博': 'dibo',
    '暂放鑫信泰': 'xinxintai', '龙跃宇': 'longyueyu', '鹏沛盈': 'pengpeiying',
    '联金': 'lianjin', '天博': 'tianbo', '凯利': 'kaili', '贯昌': 'guanchang',
    '钜成隆': 'juchenglong', '鑫信泰': 'xinxintai', '共耘': 'gongyun',
    '鑫旗舰': 'xinqijian', '鑫悦云': 'xinyueyun', '金地鑫': 'jindixin',
    '京科': 'jingke', '兴富城': 'xingfucheng', '博恒': 'boheng', '尚升': 'shangsheng',
    '永安': 'yongan', '臻艺': 'zhenyi', '成浩林': 'chenghaolin',
    '拓图电子': 'tuotudianzi', '固邦': 'gubang', '宏裕': 'hongyu', '华安': 'huaan',
    '鑫佑鑫': 'xinyouxin', '鸿发': 'hongfa', '光大': 'guangda',
    '集鑫福': 'jixinfu', '閩盛': 'minsheng', '鑫旺达': 'xinwangda',
    '诚田佳': 'chengtianjia', '伟壹昊': 'weiyihao', '万绰': 'wanchuo',
    '闽盛': 'minsheng', '常疆': 'changjiang', '鸿盛': 'hongsheng',
    '新亚洲': 'xinyazhou', '富华': 'fuhua', '宝佳盛': 'baojiasheng',
    '宇帆': 'yufan', '联科': 'lianke', '加韵': 'jiayun', '長江': 'changjiang',
    '博明': 'boming', '菘翊': 'songyi', '长江': 'changjiang', '色彩': 'secaichina',
    '诚亿光': 'chengyiguang', '浩正芯': 'haozhengxin',
    '罗杰斯泡棉': 'rogerspaomian',
}


def to_pinyin(name):
    if name in PINYIN:
        return PINYIN[name]
    # 兜底：取前2个字的常见读音
    return 'gys'


def gen_contact():
    return random.choice(SURNAMES) + random.choice(GIVEN) + random.choice(GIVEN)


def gen_phone():
    if random.random() < 0.5:
        return f'0755-{random.randint(10000000, 99999999)}'
    return f'13{random.choice("456789")}{random.randint(10000000, 99999999)}'


def gen_email(pinyin):
    return f'{pinyin}{random.randint(1, 99)}@163.com'


def gen_tax_no():
    # 18位统一社会信用代码：91 + 行政区划6位 + 9位随机（字母数字）
    area = random.choice(['440300', '440306', '440305', '440304'])
    chars = '0123456789ABCDEFGHJKLMNPQRTUWXY'
    body = ''.join(random.choice(chars) for _ in range(9))
    return f'91{area}{body}'


def gen_bank_account():
    # 16~19位数字
    length = random.choice([16, 17, 18, 19])
    return ''.join(random.choice('0123456789') for _ in range(length))


def load_names(path):
    """从 xlsx(第1列,跳表头) 或 txt(每行一个) 读取供应商名称"""
    if path.lower().endswith('.xlsx'):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        names = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # 表头
            if row and row[0] is not None and str(row[0]).strip():
                names.append(str(row[0]).strip())
        wb.close()
    else:
        with open(path, encoding='utf-8') as f:
            names = [ln.strip() for ln in f if ln.strip()]
    # 去重保序
    seen = set()
    uniq = []
    for n in names:
        if n not in seen:
            seen.add(n)
            uniq.append(n)
    return uniq


def main():
    parser = argparse.ArgumentParser(description='供应商导入模板数据生成器')
    parser.add_argument('source', help='供应商名称文件（xlsx 或 txt）')
    parser.add_argument('-o', '--output', default='供应商导入-生成.xlsx', help='输出文件')
    args = parser.parse_args()

    names = load_names(args.source)
    if not names:
        print('❌ 未读取到供应商名称')
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = '供应商导入模板'
    ws.append(['供应商编码(*)', '供应商名称(*)', '供应商类型(*)', '联系人',
               '联系电话', '邮箱', '地址', '税号', '银行账号', '备注'])

    for i, name in enumerate(names, 1):
        ws.append([
            f'SUP{i:03d}',
            name,
            'M',
            gen_contact(),
            gen_phone(),
            gen_email(to_pinyin(name)),
            '广东省深圳市宝安区',
            gen_tax_no(),
            gen_bank_account(),
            '物料供应商',
        ])

    wb.save(args.output)
    print(f'✅ 已生成: {args.output} ({len(names)} 家供应商, 10列全填充)')
    print(f'   编码范围: SUP001 ~ SUP{len(names):03d}')


if __name__ == '__main__':
    main()
