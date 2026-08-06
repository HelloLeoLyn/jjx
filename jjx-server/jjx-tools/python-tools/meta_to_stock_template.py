#!/usr/bin/env python3
"""
meta-material 盘点表 → 库存导入模板 转换工具（可复用，2026-08-06 创建）
用法:
    python3 meta_to_stock_template.py <盘点表路径> [输出路径]

功能:
    把车间物料盘点 Excel（meta-material 格式）转换成系统「库存导入模板」格式，
    供 库存列表→导入 使用。

盘点表列结构（自动识别）:
    A=材料名称  B=规格  C=上月结存  ...  BP(68)=本月结存  BQ(69)=库存平方
    BR(70)=原进料  BS(71)=盘点日期  BT(72)=备注/说明  BU(73)=摆放/区域

映射规则:
    物料名称(*) ← A 材料名称
    规格        ← B 规格
    库存数量(*) ← BP 本月结存（空则 C 上月结存）
    备注        ← BT 备注/说明
    摆放区域    ← BU 摆放/区域
    供应商      ← 材料名括号内提取（过滤数字/颜色/属性词），提取不到 → JJX（通用供应商）
    仓库/批次号/单位成本/生产日期/到期日期 ← 盘点表无此数据，留空

依赖: openpyxl（pip install openpyxl）
"""
import re
import sys
from openpyxl import load_workbook

# 供应商黑名单（材料名括号里的非供应商词：规格/颜色/属性）
BLACKLIST = set('''0.5 1.0 1.5 2.0 2.5 3.5T无胶 0603红灯 0603翠绿 0603黄灯 0805橙灯 0805白灯 0805红灯
0805绿灯 0805蓝灯 0805黄灯 3216橙绿 3216红灯 3216绿灯 3216翠绿 3216黄灯 3216黄绿 AD客供料 AG-80 AGK
AU BM BM博明 CHL DB6842 DB98KJ DSMS FR35F GF客供 KC客供 LJ PC32V RGB ROHM SDT125 SDT188 SDT25
SDT50 SDT75 SH SS TLT25白 TW TW新晟 XBQ XXT YA ZQ客供 不能贴视窗口 哑面粗砂 共阴 国外 客供 易碎
粗砂 网购 自带离型膜 白色 黑色 银亮 銀亮 高弹力 铝箔 西卡纸'''.split())

# 非物料行过滤关键词
FILTER_KW = ['下面是', '请示', '材料名称', '以下', '盘点', '评估', '处理办法', '说明', '材料']

# 列索引（1-based，对应盘点表固定结构）
COL_NAME = 1    # A 材料名称
COL_SPEC = 2    # B 规格
COL_LAST = 3    # C 上月结存
COL_BP = 68     # BP 本月结存
COL_BT = 72     # BT 备注/说明
COL_BU = 73     # BU 摆放/区域


def extract_supplier(name: str) -> str:
    """从材料名括号提取供应商，提取不到返回 JJX"""
    for m in re.findall(r'[（(]([^（）()]*)[）)]', name):
        s = m.strip()
        if s and s not in BLACKLIST and not s.isdigit() and len(s) <= 6:
            return s
    return 'JJX'


def is_valid_material(name: str) -> bool:
    """过滤非物料行"""
    n = name.strip()
    if not n or n.isdigit():
        return False
    for kw in FILTER_KW:
        if kw in n:
            return False
    return True


def main():
    if len(sys.argv) < 2:
        print("用法: python3 meta_to_stock_template.py <盘点表.xlsx> [输出.xlsx]")
        sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else '库存导入-生成.xlsx'

    wb = load_workbook(src, data_only=True)
    ws = wb.active
    print(f"读取: {src} 共 {ws.max_row} 行")

    out_rows = []
    skipped = 0
    for r in range(3, ws.max_row + 1):  # 第1行日期、第2行表头，第3行起数据
        name = ws.cell(r, COL_NAME).value
        if name is None:
            continue
        name = str(name).strip()
        if not is_valid_material(name):
            skipped += 1
            continue

        spec = ws.cell(r, COL_SPEC).value
        bp = ws.cell(r, COL_BP).value
        last = ws.cell(r, COL_LAST).value
        qty = bp if bp is not None else last

        out_rows.append({
            '物料名称': name,
            '规格': spec if spec is not None else '',
            '库存数量': qty if qty is not None else '',
            '备注': ws.cell(r, COL_BT).value or '',
            '摆放区域': ws.cell(r, COL_BU).value or '',
            '供应商': extract_supplier(name),
            '仓库': '',
            '批次号': '',
            '单位成本': '',
            '生产日期': '',
            '到期日期': '',
        })

    # 写库存导入模板格式（表头与系统模板一致，带星号由系统自动加）
    from openpyxl import Workbook
    owb = Workbook()
    ows = owb.active
    ows.title = '库存导入模板'
    headers = ['物料名称', '规格', '库存数量', '备注', '摆放区域', '供应商', '仓库', '批次号', '单位成本', '生产日期', '到期日期']
    ows.append(headers)
    for row in out_rows:
        ows.append([row[h] for h in headers])
    owb.save(out)

    # 统计
    sup_cnt = {}
    for row in out_rows:
        sup_cnt[row['供应商']] = sup_cnt.get(row['供应商'], 0) + 1
    print(f"✅ 已生成: {out} ({len(out_rows)} 行, 跳过非物料行 {skipped})")
    print(f"供应商分布: JJX={sup_cnt.get('JJX', 0)} 其他={len(sup_cnt) - 1} 个")
    for s, c in sorted(sup_cnt.items(), key=lambda x: -x[1])[:8]:
        print(f"  {s}: {c}")


if __name__ == '__main__':
    main()
