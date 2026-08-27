#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 去重工具类（可复用，2026-08-07 创建）

功能：
    按客户指定的列（1~5 列，动态）对 Excel 数据去重：
    - 支持按「列名」或「列索引」指定去重列
    - 支持同时按多列组合去重（最多 5 列）
    - 支持保留策略：保留首次出现（默认）或保留末次出现
    - 输出：去重后的 Excel + 控制台统计（原始行数/去重行数/删除行数/重复样例）

用法（命令行）:
    python3 excel_dedup.py <输入.xlsx> -o <输出.xlsx> -c 物料名称,规格
    python3 excel_dedup.py <输入.xlsx> -o <输出.xlsx> -c 0,1          # 按列索引
    python3 excel_dedup.py <输入.xlsx> -o <输出.xlsx> -c 材料,规格,供应商 -k last

用法（代码引用）:
    from excel_dedup import ExcelDedup
    dedup = ExcelDedup('输入.xlsx', columns=['材料名称', '规格'], keep='first')
    result = dedup.run('输出.xlsx')
    print(result)  # {'total': 1560, 'unique': 1520, 'removed': 40, 'duplicates': [...]}

依赖: openpyxl（pip install openpyxl）
"""
import argparse
import sys
from collections import OrderedDict
from openpyxl import load_workbook, Workbook


class ExcelDedup:
    """按指定列（最多5列）对 Excel 去重的工具类。"""

    MAX_COLUMNS = 5

    def __init__(self, source: str, columns, keep: str = 'first', sheet: str = None,
                 header_row: int = 1, skip_rows: int = 0):
        """
        :param source: 输入 Excel 文件路径
        :param columns: 去重列，list[str]（列名）或 list[int]（1-based 列索引），长度 1~5
        :param keep: 'first' 保留首次出现（默认）/'last' 保留末次出现
        :param sheet: 工作表名，默认第一个 sheet
        :param header_row: 表头所在行号（1-based），默认 1
        :param skip_rows: 表头之前额外跳过的行数（如说明行），默认 0
        """
        self.source = source
        self.columns = list(columns) if columns else []
        self.keep = keep
        self.sheet = sheet
        self.header_row = header_row
        self.skip_rows = skip_rows

        if not self.columns:
            raise ValueError('必须指定至少 1 个去重列')
        if len(self.columns) > self.MAX_COLUMNS:
            raise ValueError(f'去重列最多支持 {self.MAX_COLUMNS} 列，当前 {len(self.columns)} 列')
        if keep not in ('first', 'last'):
            raise ValueError("keep 参数只支持 'first' 或 'last'")

    # ---------- 读取 ----------
    def _load(self):
        """读取 Excel，返回 (表头列表, 数据行列表[list of cell values], 列名->列索引映射)"""
        wb = load_workbook(self.source, read_only=True, data_only=True)
        ws = wb[self.sheet] if self.sheet else wb[wb.sheetnames[0]]

        # 跳过表头前的说明行
        header_start = self.header_row + self.skip_rows
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < header_start:
            wb.close()
            raise ValueError(f'文件行数不足，无法定位表头（第 {header_start} 行）')

        header = all_rows[header_start - 1]
        # 表头去空：保留原始位置，空列名用占位
        header = [str(h).strip() if h is not None else f'__col{i+1}__' for i, h in enumerate(header)]
        data_rows = [list(r) for r in all_rows[header_start:]]
        wb.close()
        return header, data_rows

    def _resolve_columns(self, header):
        """把列名/列索引统一解析为 0-based 列索引。返回 (索引列表, 列名列表)"""
        indexes = []
        names = []
        for col in self.columns:
            if isinstance(col, int):
                idx = col - 1  # 1-based 转 0-based
            elif isinstance(col, str):
                # 兼容 "A"、"B" 字母列号
                if len(col) == 1 and col.isalpha():
                    idx = ord(col.upper()) - ord('A')
                elif col.isdigit():
                    idx = int(col) - 1
                else:
                    # 列名匹配（支持去掉 * 号的匹配，如 物料名称(*) 匹配 物料名称）
                    matched = None
                    for i, h in enumerate(header):
                        if h == col or h.replace('(*)', '').strip() == col.replace('(*)', '').strip():
                            matched = i
                            break
                    if matched is None:
                        raise ValueError(f'找不到列名: {col}，可用列: {header}')
                    idx = matched
            else:
                raise ValueError(f'不支持的列类型: {type(col)}')
            if idx < 0 or idx >= len(header):
                raise ValueError(f'列索引超出范围: {col}（文件共 {len(header)} 列）')
            indexes.append(idx)
            names.append(header[idx])
        return indexes, names

    @staticmethod
    def _cell_key(row, indexes):
        """把指定列的值拼成去重 key（None 归一为空串，数字统一去尾零）"""
        parts = []
        for idx in indexes:
            v = row[idx]
            if v is None:
                parts.append('')
            elif isinstance(v, float) and v.is_integer():
                parts.append(str(int(v)))
            else:
                parts.append(str(v).strip())
        return '|'.join(parts)

    # ---------- 主流程 ----------
    def run(self, output: str = None):
        """
        执行去重。
        :param output: 输出文件路径；None 则只统计不落盘
        :return: dict {total, unique, removed, key_columns, duplicates(重复key及行号样例)}
        """
        header, data_rows = self._load()
        col_indexes, col_names = self._resolve_columns(header)

        # 按 keep 策略决定遍历方向
        if self.keep == 'first':
            iter_rows = enumerate(data_rows)
            def should_keep(seen, key):
                if key in seen:
                    return False
                seen.add(key)
                return True
        else:
            # last：从后往前遍历，保留第一次遇到的（即原文件最后一条）
            iter_rows = reversed(list(enumerate(data_rows)))
            def should_keep(seen, key):
                if key in seen:
                    return False
                seen.add(key)
                return True

        seen = set()
        keep_rows = []      # 保留的行（原顺序输出时 first 直接顺序，last 需反转）
        removed_samples = []  # (原行号, key) 样例
        removed_count = 0
        key_stats = {}      # key -> count

        for orig_idx, row in iter_rows:
            if row is None or all(v is None or str(v).strip() == '' for v in row):
                continue  # 跳过全空行
            key = self._cell_key(row, col_indexes)
            key_stats[key] = key_stats.get(key, 0) + 1
            if should_keep(seen, key):
                keep_rows.append((orig_idx, row))
            else:
                removed_count += 1
                if len(removed_samples) < 10:
                    removed_samples.append((orig_idx + 1 + self.header_row + self.skip_rows, key))

        # last 策略下 keep_rows 是逆序收集的，恢复原顺序
        if self.keep == 'last':
            keep_rows.sort(key=lambda x: x[0])
        keep_rows = [r for _, r in keep_rows]

        dup_keys = [k for k, c in key_stats.items() if c > 1]

        result = {
            'total': len(data_rows),
            'unique': len(keep_rows),
            'removed': removed_count,
            'key_columns': col_names,
            'duplicate_keys': dup_keys[:20],
            'duplicate_count': len(dup_keys),
            'removed_samples': removed_samples,
        }

        if output:
            self._save(output, header, keep_rows)
            result['output'] = output

        return result

    def _save(self, output, header, keep_rows):
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        for row in keep_rows:
            ws.append(row)
        wb.save(output)

    # ---------- 展示 ----------
    @staticmethod
    def print_summary(result):
        print(f"去重列: {', '.join(result['key_columns'])}")
        print(f"原始行数: {result['total']}")
        print(f"去重后行数: {result['unique']}")
        print(f"删除行数: {result['removed']}")
        print(f"重复组合数: {result['duplicate_count']}")
        if result['removed_samples']:
            print("重复样例(原行号, key):")
            for r in result['removed_samples']:
                print(f"  {r}")
        if 'output' in result:
            print(f"已输出: {result['output']}")


def main():
    parser = argparse.ArgumentParser(description='Excel 去重工具（按指定列，最多5列）')
    parser.add_argument('source', help='输入 Excel 文件')
    parser.add_argument('-o', '--output', help='输出 Excel 文件（不填则只统计）')
    parser.add_argument('-c', '--columns', required=True,
                        help='去重列，逗号分隔。支持列名(物料名称,规格)或列索引(0,1)或字母(A,B)')
    parser.add_argument('-k', '--keep', choices=['first', 'last'], default='first',
                        help='保留策略: first=保留首次出现(默认), last=保留末次出现')
    parser.add_argument('-s', '--sheet', help='工作表名（默认第一个）')
    parser.add_argument('-H', '--header-row', type=int, default=1, help='表头所在行号(1-based，默认1)')
    parser.add_argument('-S', '--skip-rows', type=int, default=0, help='表头前额外跳过的行数')
    args = parser.parse_args()

    try:
        columns = [c.strip() for c in args.columns.split(',') if c.strip()]
        dedup = ExcelDedup(args.source, columns, keep=args.keep, sheet=args.sheet,
                           header_row=args.header_row, skip_rows=args.skip_rows)
        result = dedup.run(args.output)
        ExcelDedup.print_summary(result)
    except Exception as e:
        print(f'❌ 执行失败: {e}', file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
